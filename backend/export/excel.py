"""
Excel export — replicates Melissa's tracker format exactly.

Column order (all sites) follows spec F:
  1. Invoice metadata: invoice_number (or meter_number), bill_date, due_date
  2. Period: enbridge_qtr_reference, start_date, end_date, billing_period (+ rate for Walgreen)
  3. Consumption and charges (site-specific)
  4. Totals: hst_amount, total_incl_hst (Enbridge sites)
  5. Balance: balance_forward, late_payment_charge (where applicable)
  6. Calculated: cost_per_m3 or cost_per_kwh  ← Excel formula
  7. Metadata: source_pdf_filename
  8. Notes: always last

Enbridge trackers (Cambridge, Pickering CNG, Walgreen):
  Row 1 : site title
  Row 2 : Address
  Row 3 : Account Number
  Row 4 : Bill Number  (+ col B note for Cambridge/Pickering)
  Row 5 : Rate description
  Rows 6-7: empty (Walgreen row 7 has a note)
  Row 8 : column headers  (blue background, white text)
  Row 9+: data — with Excel formulas for calculated fields

Elexicon tracker:
  Row 1 : group labels ("Distribution Charges" merged O1:R1, "Other Charges" merged S1:Y1)
  Row 2 : column headers
  Row 3+: data — with Excel formulas

Cambridge formula columns (new layout):
  col 18 (R) = enbridge_invoice_cost_excl_hst: =SUM(K{r}:Q{r})
  col 20 (T) = total_incl_hst:                 =R{r}+S{r}
  col 23 (W) = cost_per_m3:                    =IF(I{r}=0,0,R{r}/I{r})

Pickering CNG formula columns:
  col 27 (AA) = total_incl_hst:  =Y{r}+Z{r}
  col 29 (AC) = cost_per_m3:     =IF(L{r}=0,0,Y{r}/L{r})

Walgreen formula columns:
  col 25 (Y) = total_incl_hst:  =W{r}+X{r}
  col 26 (Z) = cost_per_m3:     =IF(M{r}=0,0,W{r}/M{r})

Elexicon formula columns (new layout):
  col 28 (AB) = total_charge_excl_hst_interest: =AA{r}-Z{r}-IF(ISBLANK(R{r}),0,R{r})
  col 29 (AC) = cost_per_kwh:                   =IF(K{r}=0,0,AB{r}/K{r})
"""

from __future__ import annotations

import io
from datetime import date
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────
# Style constants
# ─────────────────────────────────────────────────────────────
_HDR_FILL   = PatternFill("solid", fgColor="1E40AF")   # blue-800
_HDR_FONT   = Font(bold=True, color="FFFFFF")
_GRP_FILL   = PatternFill("solid", fgColor="3B82F6")   # blue-500
_GRP_FONT   = Font(bold=True, color="FFFFFF")
_DOLLAR_FMT = '#,##0.00'
_DATE_FMT   = 'YYYY-MM-DD'
_DATE_FIELDS = {"start_date", "end_date", "billing_period", "bill_date", "due_date"}

# ─────────────────────────────────────────────────────────────
# Site metadata
# ─────────────────────────────────────────────────────────────
_ENBRIDGE_META: dict[str, dict[str, Any]] = {
    "cambridge": {
        "sheet_name": "Cambridge CNG Invoice Tracking",
        "title": "Cambridge Enbridge CNG Invoice Tracking",
        "address": "Address: 2138 EAGLEST N, CAMBRIDGE ON N3H0A1",
        "account": "Account Number: 1183991",
        "bill": "Bill Number: BA4297",
        "bill_note": "<< in Sage",
        "rate": (
            "Rate:\n"
            " - Rate M4 Firm Industrial and Commercial (Contracted Demand= 30,500.0 m3)\n"
            " - Rate M5 Interruptible"
        ),
        "row7_note": None,
    },
    "pickering_cng": {
        "sheet_name": "Enbridge Invoice Comparison",
        "title": "Enbridge Invoice - Pickering",
        "address": "Address: 1250 SQUIRES BEACH RD CNG STATION PICKERING ON L1L 1L1",
        "account": "Account Number: 930610223601",
        "bill": "Bill Number: 107000512616",
        "bill_note": "<< bill number in sage (include private check box)",
        "rate": "Rate: Rate 100",
        "row7_note": None,
    },
    "walgreen": {
        "sheet_name": "Walgreen CNG Invoice Tracking",
        "title": "Walgreen Enbridge CNG Invoice Tracking",
        "address": "Address: 145 WALGREEN ROAD, CARP, ON K0A 1L0",
        "account": "Account Number: TBD",
        "bill": "Bill Number: TBD",
        "bill_note": None,
        "rate": (
            "Rate:\n"
            " - Rate 110 Firm - CD = 10,577 m3\n"
            " - Rate 145 Interruptible - CD = 10,577 m3"
        ),
        "row7_note": "Will need to update the table below once an invoice is received for walgreen",
    },
}

# ─────────────────────────────────────────────────────────────
# Column definitions  (header, field_key, is_dollar, is_date)
# ─────────────────────────────────────────────────────────────
# Cambridge: 25 columns (A–Y)
# Col 18 (R) = enbridge_invoice_cost_excl_hst  → formula =SUM(K:Q)
# Col 20 (T) = total_incl_hst                  → formula =R+S
# Col 23 (W) = cost_per_m3                     → formula =IF(I=0,0,R/I)
_CambridgeCols = [
    ("Invoice #",                               "invoice_number",                  False, False),  # A=1
    ("Bill Date",                               "bill_date",                       False, True),   # B=2
    ("Due Date",                                "due_date",                        False, True),   # C=3
    ("Enbridge Qtr Handbook Rate Reference",    "enbridge_qtr_reference",          False, False),  # D=4
    ("Start Date",                              "start_date",                      False, True),   # E=5
    ("End Date",                                "end_date",                        False, True),   # F=6
    ("Billing Period",                          "billing_period",                  False, True),   # G=7
    ("CD (m³)",                                 "cd",                              False, False),  # H=8
    ("Gas Consumption (m³)",                    "gas_consumption",                 False, False),  # I=9
    ("Split Volumes (m³)",                      "split_volumes",                   False, False),  # J=10
    ("Demand Charge (First 8,450 m³ of CD)",    "demand_charge",                   True,  False),  # K=11
    ("Delivery Charge - First 422,250 m³",      "delivery_charge",                 True,  False),  # L=12
    ("Monthly Charge - Interruptible",          "monthly_charge_interruptible",    True,  False),  # M=13
    ("Gas Supply - Commodity",                  "gas_supply_commodity",            True,  False),  # N=14
    ("Gas Supply - Transportation",             "gas_supply_transportation",       True,  False),  # O=15
    ("Commodity & Fuel - Price Adjustment",     "commodity_fuel_price_adjustment", True,  False),  # P=16
    ("Miscellaneous Charges",                   "miscellaneous_charges",           True,  False),  # Q=17
    ("Enbridge Invoice Cost (Excluding HST)",   "enbridge_invoice_cost_excl_hst",  True,  False),  # R=18 FORMULA
    ("HST",                                     "hst_amount",                      True,  False),  # S=19
    ("Total (Incl. HST)",                       "total_incl_hst",                  True,  False),  # T=20 FORMULA
    ("Balance Forward",                         "balance_forward",                 True,  False),  # U=21
    ("Late Payment Charge",                     "late_payment_charge",             True,  False),  # V=22
    ("$/m³",                                    "cost_per_m3",                     False, False),  # W=23 FORMULA
    ("Source PDF",                              "source_pdf_filename",             False, False),  # X=24
    ("Notes",                                   "notes",                           False, False),  # Y=25
]

# Pickering CNG: 31 columns (A–AE)
# Col 27 (AA) = total_incl_hst → formula =Y+Z
# Col 29 (AC) = cost_per_m3   → formula =IF(L=0,0,Y/L)
_PickeringCngCols = [
    ("Invoice # (Bill Number)",                 "invoice_number",                  False, False),  # A=1
    ("Bill Date",                               "bill_date",                       False, True),   # B=2
    ("Due Date",                                "due_date",                        False, True),   # C=3
    ("Enbridge Qtr Handbook Rate Reference",    "enbridge_qtr_reference",          False, False),  # D=4
    ("Start Date",                              "start_date",                      False, True),   # E=5
    ("End Date",                                "end_date",                        False, True),   # F=6
    ("Billing Period",                          "billing_period",                  False, False),  # G=7
    ("Meter Reading Previous",                  "meter_reading_previous",          False, False),  # H=8
    ("Meter Reading Actual",                    "meter_reading_actual",            False, False),  # I=9
    ("CF to M3 Conversion",                     "cf_to_m3_conversion",             False, False),  # J=10
    ("CD (m³)",                                 "cd",                              False, False),  # K=11
    ("Gas Consumption (m³)",                    "gas_consumption",                 False, False),  # L=12
    ("Split Volumes (m³)",                      "split_volumes",                   False, False),  # M=13
    ("Customer Charge",                         "customer_charge",                 True,  False),  # N=14
    ("CD _1",                                   "cd_1",                            True,  False),  # O=15
    ("CD_2",                                    "cd_2",                            True,  False),  # P=16
    ("Delivery To You",                         "delivery_to_you",                 True,  False),  # Q=17
    ("Load Balancing",                          "load_balancing",                  True,  False),  # R=18
    ("Transportation",                          "transportation",                  True,  False),  # S=19
    ("Federal Carbon Charge",                   "federal_carbon_charge",           True,  False),  # T=20
    ("Gas Supply Charge_1",                     "gas_supply_charge_1",             True,  False),  # U=21
    ("Gas Supply Charge_2",                     "gas_supply_charge_2",             True,  False),  # V=22
    ("Cost Adjustment",                         "cost_adjustment",                 True,  False),  # W=23
    ("Previous Bill charge",                    "previous_bill_charge",            True,  False),  # X=24
    ("Enbridge Invoice Cost (Excluding HST)",   "enbridge_invoice_cost_excl_hst",  True,  False),  # Y=25
    ("HST",                                     "hst_amount",                      True,  False),  # Z=26
    ("Total (Incl. HST)",                       "total_incl_hst",                  True,  False),  # AA=27 FORMULA
    ("Balance Forward",                         "balance_forward",                 True,  False),  # AB=28
    ("$/m³",                                    "cost_per_m3",                     False, False),  # AC=29 FORMULA
    ("Source PDF",                              "source_pdf_filename",             False, False),  # AD=30
    ("Notes",                                   "notes",                           False, False),  # AE=31
]

# Walgreen: 28 columns (A–AB)
# Col 25 (Y) = total_incl_hst → formula =W+X
# Col 26 (Z) = cost_per_m3   → formula =IF(M=0,0,W/M)
_WalgreenCols = [
    ("Invoice # (Bill Number)",                 "invoice_number",                  False, False),  # A=1
    ("Bill Date",                               "bill_date",                       False, True),   # B=2
    ("Due Date",                                "due_date",                        False, True),   # C=3
    ("Enbridge Qtr Handbook Rate Reference",    "enbridge_qtr_reference",          False, False),  # D=4
    ("Rate",                                    "rate",                            False, False),  # E=5
    ("Start Date",                              "start_date",                      False, True),   # F=6
    ("End Date",                                "end_date",                        False, True),   # G=7
    ("Days",                                    "days",                            False, False),  # H=8
    ("CD_1 (m³)",                               "cd_1",                            False, False),  # I=9
    ("CD_2 (m³)",                               "cd_2",                            False, False),  # J=10
    ("Gas Consumption_1 (m³)",                  "gas_consumption_1",               False, False),  # K=11
    ("Gas Consumption_2 (m³)",                  "gas_consumption_2",               False, False),  # L=12
    ("Total Gas Consumption (m³)",              "total_gas_consumption",           False, False),  # M=13
    ("Customer Monthly Charge",                 "customer_monthly_charge",         True,  False),  # N=14
    ("Demand Charge",                           "demand_charge",                   True,  False),  # O=15
    ("Demand Charge 2",                         "demand_charge_2",                 True,  False),  # P=16
    ("Delivery Charge",                         "delivery_charge",                 True,  False),  # Q=17
    ("Loading Balance Charge",                  "load_balancing_charge",           True,  False),  # R=18
    ("Transportation",                          "transportation",                  True,  False),  # S=19
    ("Gas Supply - Commodity",                  "gas_supply_commodity",            True,  False),  # T=20
    ("Gas Supply - Commodity_2",                "gas_supply_commodity_2",          True,  False),  # U=21
    ("Cost Adjustment",                         "cost_adjustment",                 True,  False),  # V=22
    ("Enbridge Invoice Cost (Excluding HST)",   "enbridge_invoice_cost_excl_hst",  True,  False),  # W=23
    ("HST",                                     "hst_amount",                      True,  False),  # X=24
    ("Total (Incl. HST)",                       "total_incl_hst",                  True,  False),  # Y=25 FORMULA
    ("$/m³",                                    "cost_per_m3",                     False, False),  # Z=26 FORMULA
    ("Source PDF",                              "source_pdf_filename",             False, False),  # AA=27
    ("Notes",                                   "notes",                           False, False),  # AB=28
]

# Elexicon: 31 columns (A–AE)
# Distribution Charges group: cols 15-18 (O–R) → merge O1:R1
# Other Charges group: cols 19-25 (S–Y) → merge S1:Y1
# Col 28 (AB) = total_charge_excl_hst_interest → formula =AA-Z-IF(ISBLANK(R),0,R)
# Col 29 (AC) = cost_per_kwh                   → formula =IF(K=0,0,AB/K)
_ElexiconCols = [
    ("Meter Number",                                              "meter_number",                  False, False),  # A=1
    ("Bill Date",                                                 "bill_date",                     False, True),   # B=2
    ("Due Date",                                                  "due_date",                      False, True),   # C=3
    ("Bill Period",                                               "bill_period",                   False, False),  # D=4
    ("Read Period",                                               "read_period",                   False, False),  # E=5
    ("Start Date",                                                "start_date",                    False, True),   # F=6
    ("End Date",                                                  "end_date",                      False, True),   # G=7
    ("Account Number",                                            "account_number",                False, False),  # H=8
    ("Service Type",                                              "service_type",                  False, False),  # I=9
    ("Days",                                                      "days",                          False, False),  # J=10
    ("kWh Used",                                                  "kwh_used",                      False, False),  # K=11
    ("Monthly Demand (kW)",                                       "monthly_demand_kw",             False, False),  # L=12
    ("Electricity ($/kWh)",                                       "electricity_rate",              False, False),  # M=13
    ("Global Adjuster ($)",                                       "global_adjuster",               False, False),  # N=14
    # ← Distribution Charges group: O1:R1 (cols 15–18)
    ("New Account Setup",                                         "new_account_setup",             True,  False),  # O=15
    ("Delivery Charge",                                           "delivery_charge",               True,  False),  # P=16
    ("Customer Charge",                                           "customer_charge",               True,  False),  # Q=17
    ("Interest Overdue Charge",                                   "interest_overdue_charge",       True,  False),  # R=18
    # ← Other Charges group: S1:Y1 (cols 19–25)
    ("SSS admin charge",                                          "sss_admin_charge",              True,  False),  # S=19
    ("Electriciy",                                                "electricity_cost",              True,  False),  # T=20  (tracker typo preserved)
    ("Global Adjustment",                                         "global_adjustment",             True,  False),  # U=21
    ("Global Adjustment Recovery",                                "global_adjustment_recovery",    True,  False),  # V=22
    ("Transmission Network Charge",                               "transmission_network",          True,  False),  # W=23
    ("Transmission Connection",                                   "transmission_connection",       True,  False),  # X=24
    ("Wholesale Market Services",                                 "wholesale_market_services",     True,  False),  # Y=25
    ("H.S.T",                                                     "hst",                           True,  False),  # Z=26
    ("Total Charge",                                              "total_charge",                  True,  False),  # AA=27
    ("Total Charge (Excluding HST & Overdue Interest Charges)",   "total_charge_excl_hst_interest", True, False),  # AB=28 FORMULA
    ("$/kWh",                                                     "cost_per_kwh",                  False, False),  # AC=29 FORMULA
    ("Source PDF",                                                "source_pdf_filename",           False, False),  # AD=30
    ("Notes",                                                     "notes",                         False, False),  # AE=31
]

_SITE_COLS = {
    "cambridge":          _CambridgeCols,
    "pickering_cng":      _PickeringCngCols,
    "walgreen":           _WalgreenCols,
    "pickering_elexicon": _ElexiconCols,
}


# ─────────────────────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────────────────────

def export_to_excel(site_id: str, records: list[dict[str, Any]]) -> bytes:
    """Return raw bytes of an .xlsx workbook matching Melissa's tracker format."""
    wb = Workbook()
    ws = wb.active

    if site_id == "pickering_elexicon":
        _write_elexicon(ws, records)
    else:
        _write_enbridge(ws, site_id, records)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────

def _coerce(field: str, val: Any) -> Any:
    """Convert ISO date strings to date objects so Excel formats them correctly."""
    if val is None:
        return None
    if field in _DATE_FIELDS and isinstance(val, str):
        try:
            return date.fromisoformat(val)
        except ValueError:
            return val
    return val


def _style_header_cell(cell, fill=None, font=None):
    cell.fill = fill or _HDR_FILL
    cell.font = font or _HDR_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_enbridge(ws, site_id: str, records: list[dict[str, Any]]) -> None:
    meta = _ENBRIDGE_META[site_id]
    cols = _SITE_COLS[site_id]
    ws.title = meta["sheet_name"]

    # ── Rows 1–5: site metadata ──────────────────────────────────
    ws.cell(row=1, column=1, value=meta["title"])
    ws.cell(row=2, column=1, value=meta["address"])
    ws.cell(row=3, column=1, value=meta["account"])
    ws.cell(row=4, column=1, value=meta["bill"])
    if meta["bill_note"]:
        ws.cell(row=4, column=2, value=meta["bill_note"])
    ws.cell(row=5, column=1, value=meta["rate"])
    if meta["row7_note"]:
        ws.cell(row=7, column=1, value=meta["row7_note"])

    # ── Row 8: column headers ────────────────────────────────────
    for col_i, (header, _, _, _) in enumerate(cols, 1):
        cell = ws.cell(row=8, column=col_i, value=header)
        _style_header_cell(cell)

    # ── Rows 9+: data with formulas ──────────────────────────────
    data_start = 9
    for row_offset, record in enumerate(records):
        data_row = data_start + row_offset
        for col_i, (_, field, is_dollar, is_date) in enumerate(cols, 1):
            cell = ws.cell(row=data_row, column=col_i)

            formula = _enbridge_formula(site_id, col_i, data_row)
            if formula:
                cell.value = formula
                if is_dollar:
                    cell.number_format = _DOLLAR_FMT
                continue

            val = _coerce(field, record.get(field))
            cell.value = val
            if is_dollar and val is not None:
                cell.number_format = _DOLLAR_FMT
            elif is_date and val is not None:
                cell.number_format = _DATE_FMT


def _enbridge_formula(site_id: str, col_i: int, row: int) -> str | None:
    """Return an Excel formula string for calculated columns, or None for regular cells."""
    if site_id == "cambridge":
        # Col 18 (R): =SUM(K{r}:Q{r}) — enbridge_invoice_cost_excl_hst
        if col_i == 18:
            return f"=SUM(K{row}:Q{row})"
        # Col 20 (T): =R{r}+S{r} — total_incl_hst
        if col_i == 20:
            return f"=R{row}+S{row}"
        # Col 23 (W): =IF(I{r}=0,0,R{r}/I{r}) — cost_per_m3
        if col_i == 23:
            return f"=IF(I{row}=0,0,R{row}/I{row})"

    elif site_id == "pickering_cng":
        # Col 27 (AA): =Y{r}+Z{r} — total_incl_hst
        if col_i == 27:
            return f"=Y{row}+Z{row}"
        # Col 29 (AC): =IF(L{r}=0,0,Y{r}/L{r}) — cost_per_m3
        if col_i == 29:
            return f"=IF(L{row}=0,0,Y{row}/L{row})"

    elif site_id == "walgreen":
        # Col 25 (Y): =W{r}+X{r} — total_incl_hst
        if col_i == 25:
            return f"=W{row}+X{row}"
        # Col 26 (Z): =IF(M{r}=0,0,W{r}/M{r}) — cost_per_m3
        if col_i == 26:
            return f"=IF(M{row}=0,0,W{row}/M{row})"

    return None


def _write_elexicon(ws, records: list[dict[str, Any]]) -> None:
    cols = _ElexiconCols
    ws.title = "Past Elexicon Bill"
    data_start = 3

    # ── Row 1: group header cells (merged spans) ─────────────────
    # Distribution Charges: cols 15–18 (O1:R1)
    ws.merge_cells("O1:R1")
    cell_dc = ws["O1"]
    cell_dc.value = "Distribution Charges"
    _style_header_cell(cell_dc, fill=_GRP_FILL, font=_GRP_FONT)

    # Other Charges: cols 19–25 (S1:Y1)
    ws.merge_cells("S1:Y1")
    cell_oc = ws["S1"]
    cell_oc.value = "Other Charges"
    _style_header_cell(cell_oc, fill=_GRP_FILL, font=_GRP_FONT)

    # ── Row 2: column headers ────────────────────────────────────
    for col_i, (header, _, _, _) in enumerate(cols, 1):
        cell = ws.cell(row=2, column=col_i, value=header)
        _style_header_cell(cell)

    # ── Rows 3+: data with formulas ──────────────────────────────
    for row_offset, record in enumerate(records):
        data_row = data_start + row_offset
        for col_i, (_, field, is_dollar, is_date) in enumerate(cols, 1):
            cell = ws.cell(row=data_row, column=col_i)

            formula = _elexicon_formula(col_i, data_row)
            if formula:
                cell.value = formula
                if is_dollar:
                    cell.number_format = _DOLLAR_FMT
                continue

            val = _coerce(field, record.get(field))
            cell.value = val
            if is_dollar and val is not None:
                cell.number_format = _DOLLAR_FMT
            elif is_date and val is not None:
                cell.number_format = _DATE_FMT


def _elexicon_formula(col_i: int, row: int) -> str | None:
    """Return Excel formula for Elexicon calculated columns, or None."""
    # Col 28 (AB): total_charge_excl_hst_interest = total_charge(AA) - hst(Z) - interest(R)
    if col_i == 28:
        return f"=AA{row}-Z{row}-IF(ISBLANK(R{row}),0,R{row})"
    # Col 29 (AC): cost_per_kwh = total_excl(AB) / kwh_used(K)
    if col_i == 29:
        return f"=IF(K{row}=0,0,AB{row}/K{row})"
    return None
